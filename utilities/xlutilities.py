import openpyxl
from openpyxl.styles import PatternFill


def max_rows(filepath,sheetname):
    workbook=openpyxl.load_workbook(filepath,'r')
    sheet=workbook[sheetname]
    return sheet.max_row

def max_columns(filepath,sheetname):
    workbook=openpyxl.load_workbook(filepath)
    sheet=workbook[sheetname]
    return sheet.max_column

def read_data(filepath,sheetname,r,c):
    workbook=openpyxl.load_workbook(filepath)
    sheet=workbook[sheetname]
    return sheet.cell(row=r,column=c).value

def write_data(filepath,sheetname,r,c,data):
    workbook=openpyxl.load_workbook(filepath)
    sheet=workbook[sheetname]
    sheet.cell(row=r,column=c).value=data
    workbook.save(filepath)

def fill_green(filepath,sheetname,r,c):
    workbook=openpyxl.load_workbook(filepath)
    sheet=workbook[sheetname]
    greenfill=PatternFill(start_color='60b202',end_color='60b202',fill_type='solid')
    sheet.cell(r,c).fill=greenfill
    workbook.save(filepath)

def fill_red(filepath,sheetname,r,c):
    workbook=openpyxl.load_workbook(filepath)
    sheet=workbook[sheetname]
    redfill=PatternFill(start_color='ff0000',end_color='ff0000',fill_type='solid')
    sheet.cell(row=r,column=c).fill=redfill
    workbook.save(filepath)
